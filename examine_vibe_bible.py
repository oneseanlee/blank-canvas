import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Disable pandas truncation
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', None)
pd.set_option('display.width', None)

print("=" * 80)
print("EXAMINING VIBE CODING BIBLE")
print("=" * 80)

# Load the Excel file
file_path = '/home/ubuntu/Uploads/Vibe Coding Bible.xlsx'

# First, check what sheets exist
wb = load_workbook(file_path)
sheet_names = wb.sheetnames

print(f"\n📊 WORKBOOK STRUCTURE")
print("-" * 80)
print(f"Total Sheets/Tabs: {len(sheet_names)}")
print(f"\nSheet Names:")
for i, name in enumerate(sheet_names, 1):
    print(f"  {i}. {name}")

# Load all sheets
all_sheets = pd.read_excel(file_path, sheet_name=None)

print("\n" + "=" * 80)
print("DETAILED TAB ANALYSIS")
print("=" * 80)

total_prompts = 0
tab_details = []

for sheet_name, df in all_sheets.items():
    print(f"\n📑 TAB: {sheet_name}")
    print("-" * 80)
    print(f"Rows: {len(df)}")
    print(f"Columns: {len(df.columns)}")
    print(f"Column Names: {list(df.columns)}")
    
    total_prompts += len(df)
    tab_details.append({
        'tab': sheet_name,
        'rows': len(df),
        'columns': list(df.columns)
    })
    
    # Show first 3 rows as sample
    if len(df) > 0:
        print(f"\n📋 Sample Data (First 3 rows):")
        print("-" * 80)
        for idx, row in df.head(3).iterrows():
            print(f"\nRow {idx + 1}:")
            for col in df.columns:
                value = str(row[col])[:100] if pd.notna(row[col]) else "N/A"
                print(f"  {col}: {value}")

print("\n" + "=" * 80)
print("SUMMARY STATISTICS")
print("=" * 80)
print(f"Total Tabs: {len(sheet_names)}")
print(f"Total Prompts/Rows: {total_prompts}")

# Check for categories if they exist
for sheet_name, df in all_sheets.items():
    if 'Category' in df.columns:
        categories = df['Category'].value_counts()
        print(f"\n📊 Categories in '{sheet_name}':")
        print(categories.head(10))

wb.close()

# Save summary
import json
summary = {
    'total_tabs': len(sheet_names),
    'total_prompts': total_prompts,
    'tab_details': tab_details
}

with open('/home/ubuntu/vibe_bible_summary.json', 'w') as f:
    json.dump(summary, f, indent=2)

print("\n✓ Summary saved to ~/vibe_bible_summary.json")

