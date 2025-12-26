import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load the workbook
wb = load_workbook('/home/ubuntu/vibe_coding_prompts_library_organized.xlsx')

# Get all sheets except the guide
sheet_names = [s for s in wb.sheetnames if '📚' not in s and '📋 Other' not in s]
print(f"Content sheets: {len(sheet_names)}")

# Count prompts per sheet
sheets_info = []
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    count = ws.max_row - 1  # Exclude header
    sheets_info.append({'name': sheet_name, 'count': count})
    print(f"  {sheet_name}: {count} prompts")

total_prompts = sum(s['count'] for s in sheets_info)

# Remove old guide if exists
if "📚 Quick Reference Guide" in wb.sheetnames:
    del wb["📚 Quick Reference Guide"]

# Create new guide
ws_guide = wb.create_sheet(title="📚 Quick Reference Guide")

guide_content = [
    ["🎯 VIBE CODING PROMPTS LIBRARY - QUICK REFERENCE GUIDE", ""],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["HOW TO USE THIS WORKBOOK", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["STEP 1: Start Your Session Right", ""],
    ["   📍 Go to: 🎯 START HERE - Pre-Session Setup", "Always begin by copying these essential guardrail prompts"],
    ["", "These prompts ensure the AI:"],
    ["", "   • Makes only the changes you request"],
    ["", "   • Maintains code quality and consistency"],
    ["", "   • Maximizes your credits by reducing back-and-forth"],
    ["", ""],
    ["STEP 2: Choose Your Design Category", ""],
    ["", "Navigate to the tab that matches your project:"],
    ["", ""],
]

# Add tab descriptions
for sheet_info in sheets_info:
    name = sheet_info['name']
    count = sheet_info['count']
    guide_content.append([f"   {name}", f"({count} prompts)"])

guide_content.extend([
    ["", ""],
    ["STEP 3: Filter and Find", ""],
    ["", "Use Excel filters to find the perfect prompt:"],
    ["", "   • Filter by Tool Compatibility (Lovable, Replit, ChatGPT)"],
    ["", "   • Filter by Prompt Type (Training Wheels, No Training Wheels)"],
    ["", ""],
    ["STEP 4: Customize and Use", ""],
    ["", "1. Copy the prompt from Column B"],
    ["", "2. Replace [placeholders] with your details"],
    ["", "3. Paste into your AI tool"],
    ["", ""],
    ["STEP 5: Combine for Best Results", ""],
    ["", "Layer prompts from multiple tabs for complex projects"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["UNDERSTANDING PROMPT TYPES", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎓 Training Wheels", "Detailed step-by-step guidance - Best for learning"],
    ["⚡ No Training Wheels", "Concise expert prompts - Best for speed"],
    ["🎨 Design", "Visual design and UI/UX focused"],
    ["📊 Strategy", "High-level planning and business strategy"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["PRO TIPS FOR SUCCESS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["✅ ALWAYS start with START HERE prompts", "Saves credits and prevents unwanted changes"],
    ["✅ Read Description/Notes column", "Contains valuable context and tips"],
    ["✅ Test iteratively", "Start simple, add complexity gradually"],
    ["✅ Combine design prompts", "Layer elements for award-winning results"],
    ["✅ Check tool compatibility", "Ensure prompt works with your AI tool"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["COMMON USE CASES & RECOMMENDED TABS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎯 SaaS Landing Page", "START HERE → Award-Winning Designs → Landing Pages → Social Proof"],
    ["🛒 E-commerce Store", "START HERE → E-commerce → UI/UX → Interactive → Performance"],
    ["📊 Dashboard/Admin", "START HERE → Dashboard & Admin → UI/UX → Performance"],
    ["🎨 Portfolio/Agency", "START HERE → Award-Winning Designs → Visual Design → Interactive"],
    ["💰 Sales Funnel", "START HERE → Conversion & Funnel → Landing Pages → Content"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["WORKBOOK STATISTICS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["Total Prompts", str(total_prompts)],
    ["Total Content Tabs", str(len(sheet_names))],
    ["Last Updated", "October 21, 2025"],
    ["", ""],
    ["", ""],
    ["🎉 Ready to build something amazing?", "Start with 🎯 START HERE and let's create!"],
])

# Write content
for row_idx, (col1, col2) in enumerate(guide_content, 1):
    ws_guide.cell(row=row_idx, column=1, value=col1)
    ws_guide.cell(row=row_idx, column=2, value=col2)
    
    # Format headers
    if '═══' in col1 or any(col1.startswith(x) for x in ['HOW TO USE', 'UNDERSTANDING', 
                                                          'PRO TIPS', 'COMMON USE', 
                                                          'WORKBOOK STATISTICS', '🎯 VIBE']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color='2C4E8C')
    
    # Format steps
    if any(col1.startswith(x) for x in ['STEP ']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=11, color='1F4788')
    
    # Format items
    if any(col1.startswith(x) for x in ['🎓', '⚡', '🎨', '📊', '✅', '🎯', '🛒', '💰']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=10)
    
    ws_guide.cell(row=row_idx, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws_guide.cell(row=row_idx, column=2).alignment = Alignment(vertical='top', wrap_text=True)

ws_guide.column_dimensions['A'].width = 45
ws_guide.column_dimensions['B'].width = 75

# Save
wb.save('/home/ubuntu/vibe_coding_prompts_library_organized.xlsx')
print(f"\n✅ Quick Reference Guide created successfully!")
print(f"✅ Total prompts: {total_prompts}")
print(f"✅ Total tabs: {len(sheet_names) + 1} (including guide)")

