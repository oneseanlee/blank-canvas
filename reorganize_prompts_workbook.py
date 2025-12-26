import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json

print("=" * 70)
print("STEP 1: LOADING AND ANALYZING CURRENT DATA")
print("=" * 70)

# Load current Excel file
df = pd.read_excel('/home/ubuntu/vibe_coding_prompts_library.xlsx')
print(f"✓ Loaded {len(df)} prompts from current library")

# Analyze categories
print(f"\n📊 Total unique categories: {df['Category'].nunique()}")
print(f"📝 Prompt types: {df['Prompt Type'].unique()}")

# Display category distribution
category_counts = df['Category'].value_counts()
print(f"\nTop 20 categories by count:")
for i, (cat, count) in enumerate(category_counts.head(20).items(), 1):
    print(f"  {i:2d}. {cat:45s} - {count:3d} prompts")

print("\n" + "=" * 70)
print("STEP 2: IDENTIFYING PROMPTS FOR EACH TAB")
print("=" * 70)

# Define tab structure and categorization
tab_structure = {
    "🎯 START HERE - Pre-Session Setup": {
        "description": "Essential guardrail prompts to use before starting any vibe coding session",
        "keywords": ["General Best Practices", "Meta Prompting"],
        "prompt_types": ["No Training Wheels"],
        "custom_filter": lambda row: any(keyword in str(row['Use Case']).lower() for keyword in 
                                       ["only make", "exact changes", "consistency", "guardrail", 
                                        "prevent", "ensure", "follow", "maintain", "do not modify"])
    },
    
    "✨ Award-Winning Homepage Designs": {
        "description": "Modern, futuristic homepage designs following award-winning patterns",
        "keywords": ["Visual Hierarchy & Layout", "Hero Sections", "Award-Winning Design"],
        "use_case_keywords": ["homepage", "hero", "above-fold", "f-pattern", "z-pattern", 
                             "attention ratio", "awwwards", "award"]
    },
    
    "🎨 Visual Design & Modern Aesthetics": {
        "description": "Color psychology, typography, and cutting-edge visual design trends",
        "keywords": ["Visual Design", "Typography", "Color Psychology", "Brand Identity"],
        "use_case_keywords": ["visual", "color", "typography", "glassmorphism", "neumorphism", 
                             "gradient", "aesthetic", "modern", "futuristic", "kinetic"]
    },
    
    "🚀 Interactive & Immersive Elements": {
        "description": "3D effects, animations, parallax, and interactive experiences",
        "keywords": ["Interactive & Immersive", "Animation & Interactivity", "Animation & Data"],
        "use_case_keywords": ["interactive", "animation", "parallax", "scroll", "3d", "immersive",
                             "micro-interaction", "hover", "gsap", "lottie"]
    },
    
    "💎 UI/UX Excellence": {
        "description": "Navigation, user experience, accessibility, and interface design",
        "keywords": ["UI/UX Design", "Navigation", "Accessibility", "UX/UI Strategy", 
                    "Responsive", "Mobile"],
        "use_case_keywords": ["navigation", "menu", "ux", "ui", "user experience", "accessibility",
                             "responsive", "mobile", "interface"]
    },
    
    "🎯 Conversion & Funnel Optimization": {
        "description": "Sales funnels, conversion optimization, and funnel strategies",
        "keywords": ["Funnel Design", "Funnel Optimization", "Conversion Optimization", 
                    "Conversion & Comparison", "Conversion & Urgency"],
        "use_case_keywords": ["funnel", "conversion", "optimize", "cro", "a/b test"]
    },
    
    "📄 Landing Pages & Lead Generation": {
        "description": "High-converting landing pages and lead capture strategies",
        "keywords": ["Landing Pages", "Landing Page Design", "Lead Generation", 
                    "Lead Generation & Conversion"],
        "use_case_keywords": ["landing page", "lead", "capture", "opt-in", "lead magnet"]
    },
    
    "💰 E-commerce & Product Pages": {
        "description": "Product showcases, shopping experiences, and e-commerce optimization",
        "keywords": ["E-commerce", "E-commerce & Conversion", "Interactive & E-commerce"],
        "use_case_keywords": ["product", "shop", "cart", "checkout", "e-commerce", "ecommerce"]
    },
    
    "📊 Dashboard & Admin Panels": {
        "description": "Data visualization, admin interfaces, and dashboard design",
        "keywords": ["Dashboard & Admin Panels", "Data Visualization"],
        "use_case_keywords": ["dashboard", "admin", "panel", "data visualization", "metrics"]
    },
    
    "🛠️ Feature Development": {
        "description": "Building specific features, components, and functionality",
        "keywords": ["Feature Development", "Development"],
        "use_case_keywords": ["feature", "component", "build", "implement", "create"]
    },
    
    "📱 Social Proof & Trust Building": {
        "description": "Testimonials, reviews, trust badges, and credibility elements",
        "keywords": ["Social Proof", "Trust & Credibility", "Trust Building", "Reputation"],
        "use_case_keywords": ["testimonial", "review", "trust", "social proof", "credibility"]
    },
    
    "📧 Content & Marketing": {
        "description": "Content strategy, email marketing, and marketing campaigns",
        "keywords": ["Content Marketing", "Email Marketing", "Campaign Strategy"],
        "use_case_keywords": ["content", "email", "marketing", "campaign", "newsletter"]
    },
    
    "🔗 API & Integration": {
        "description": "API integration, third-party services, and external connections",
        "keywords": ["API Integration", "Database Integration"],
        "use_case_keywords": ["api", "integration", "database", "third-party"]
    },
    
    "⚡ Performance & Technical": {
        "description": "Performance optimization, security, and technical improvements",
        "keywords": ["Performance", "Security", "Optimization", "Performance & UX"],
        "use_case_keywords": ["performance", "speed", "security", "optimize", "technical"]
    },
    
    "🔍 SEO & Analytics": {
        "description": "Search optimization, tracking, and analytics implementation",
        "keywords": ["SEO", "Analytics & Tracking"],
        "use_case_keywords": ["seo", "analytics", "tracking", "search"]
    },
    
    "🧪 Testing & Debugging": {
        "description": "Testing strategies, debugging, and quality assurance",
        "keywords": ["Testing", "Debugging"],
        "use_case_keywords": ["test", "debug", "qa", "quality"]
    },
    
    "📋 Other Specialized Prompts": {
        "description": "Miscellaneous prompts that don't fit other categories",
        "keywords": [],  # Will catch remaining prompts
        "use_case_keywords": []
    }
}

# Function to categorize prompts into tabs
def assign_to_tab(row):
    """Assign a prompt to the appropriate tab"""
    category = str(row['Category'])
    use_case = str(row['Use Case']).lower()
    prompt_text = str(row['Prompt']).lower()
    prompt_type = str(row['Prompt Type'])
    
    for tab_name, criteria in tab_structure.items():
        # Check for custom filter (for START HERE tab)
        if 'custom_filter' in criteria:
            if criteria['custom_filter'](row):
                return tab_name
        
        # Check category keywords
        if 'keywords' in criteria:
            for keyword in criteria['keywords']:
                if keyword.lower() in category.lower():
                    return tab_name
        
        # Check use case keywords
        if 'use_case_keywords' in criteria:
            for keyword in criteria['use_case_keywords']:
                if keyword in use_case or keyword in prompt_text:
                    return tab_name
        
        # Check prompt types for START HERE
        if 'prompt_types' in criteria:
            if prompt_type in criteria['prompt_types'] and any(
                kw in use_case for kw in ["best practice", "guardrail", "consistency"]
            ):
                return tab_name
    
    # Default to "Other Specialized Prompts" if no match
    return "📋 Other Specialized Prompts"

# Assign tabs to all prompts
df['Tab'] = df.apply(assign_to_tab, axis=1)

# Count prompts per tab
tab_counts = df['Tab'].value_counts()
print("\n📑 Prompts assigned to tabs:")
for tab_name in tab_structure.keys():
    count = tab_counts.get(tab_name, 0)
    print(f"  • {tab_name:50s} - {count:3d} prompts")

# Save tab assignments for verification
df.to_csv('/home/ubuntu/tab_assignments_debug.csv', index=False)
print(f"\n✓ Saved tab assignments to ~/tab_assignments_debug.csv for review")

print("\n" + "=" * 70)
print("STEP 3: CREATING MULTI-TAB WORKBOOK")
print("=" * 70)

# Create new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Column widths
column_widths = {
    'A': 30,  # Use Case
    'B': 60,  # Prompt
    'C': 20,  # Category
    'D': 25,  # Tool Compatibility
    'E': 20,  # Prompt Type
    'F': 40   # Description/Notes
}

def format_worksheet(ws, df_tab):
    """Apply formatting to a worksheet"""
    # Write headers
    headers = ['Use Case', 'Prompt', 'Category', 'Tool Compatibility', 'Prompt Type', 'Description/Notes']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data
    for row_idx, (_, row) in enumerate(df_tab.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row['Use Case'])
        ws.cell(row=row_idx, column=2, value=row['Prompt'])
        ws.cell(row=row_idx, column=3, value=row['Category'])
        ws.cell(row=row_idx, column=4, value=row['Tool Compatibility'])
        ws.cell(row=row_idx, column=5, value=row['Prompt Type'])
        ws.cell(row=row_idx, column=6, value=row['Description/Notes'])
    
    # Apply alternating row colors and text wrapping
    for row_idx in range(2, ws.max_row + 1):
        fill = gray_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set header row height
    ws.row_dimensions[1].height = 30
    
    # Enable filters
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

# Create sheets for each tab in order
tab_order = list(tab_structure.keys())
sheets_created = []

for tab_name in tab_order:
    df_tab = df[df['Tab'] == tab_name].copy()
    
    if len(df_tab) > 0:
        # Clean tab name for Excel (max 31 chars, remove special chars)
        sheet_name = tab_name[:31].replace('/', '-')
        ws = wb.create_sheet(title=sheet_name)
        
        # Drop the 'Tab' column before writing
        df_tab = df_tab.drop('Tab', axis=1)
        
        # Format worksheet
        format_worksheet(ws, df_tab)
        
        sheets_created.append({
            'name': tab_name,
            'sheet_name': sheet_name,
            'count': len(df_tab)
        })
        print(f"  ✓ Created sheet: {sheet_name} ({len(df_tab)} prompts)")

print("\n" + "=" * 70)
print("STEP 4: CREATING QUICK REFERENCE GUIDE")
print("=" * 70)

# Create Quick Reference Guide sheet
ws_guide = wb.create_sheet(title="📚 Quick Reference Guide")

# Add content to guide
guide_content = [
    ["🎯 VIBE CODING PROMPTS LIBRARY - QUICK REFERENCE GUIDE", ""],
    ["", ""],
    ["HOW TO USE THIS WORKBOOK", ""],
    ["", ""],
    ["1. Start with the 🎯 START HERE tab", "Copy and paste those prompts at the beginning of your vibe coding session"],
    ["", "This ensures the AI follows your instructions precisely and maximizes your credits"],
    ["", ""],
    ["2. Choose your design focus", "Navigate to the relevant tab based on what you're building:"],
    ["", "• ✨ Award-Winning Homepage Designs - For impressive, modern homepage layouts"],
    ["", "• 🎨 Visual Design & Modern Aesthetics - For color, typography, and visual styling"],
    ["", "• 🚀 Interactive & Immersive Elements - For animations and interactive features"],
    ["", "• 💎 UI/UX Excellence - For navigation and user experience"],
    ["", ""],
    ["3. Filter and search", "Use the filter dropdowns in each sheet to find specific:"],
    ["", "• Tool compatibility (Lovable, Replit, ChatGPT, etc.)"],
    ["", "• Prompt types (Training Wheels vs No Training Wheels)"],
    ["", "• Specific use cases"],
    ["", ""],
    ["4. Copy the prompt", "Copy the entire prompt text from column B"],
    ["", "Customize it by replacing [placeholders] with your specific needs"],
    ["", ""],
    ["5. Combine prompts", "For complex projects, combine prompts from multiple tabs:"],
    ["", "• Start with START HERE guardrails"],
    ["", "• Add design prompts from Visual Design tab"],
    ["", "• Add conversion elements from Landing Pages tab"],
    ["", "• Add technical requirements from Performance tab"],
    ["", ""],
    ["UNDERSTANDING PROMPT TYPES", ""],
    ["", ""],
    ["Training Wheels", "Detailed, step-by-step prompts with extensive guidance"],
    ["", "Use when: You're learning or need detailed control"],
    ["", ""],
    ["No Training Wheels", "Concise, expert-level prompts that assume knowledge"],
    ["", "Use when: You're experienced and want quick results"],
    ["", ""],
    ["Design", "Visual design and UI/UX focused prompts"],
    ["", "Use when: Creating interfaces, layouts, or visual elements"],
    ["", ""],
    ["Strategy", "High-level planning and strategic prompts"],
    ["", "Use when: Planning funnels, campaigns, or business strategies"],
    ["", ""],
    ["TOOL COMPATIBILITY GUIDE", ""],
    ["", ""],
    ["Most Compatible Tools:", ""],
    ["• Replit", "Best for full-stack applications (116 prompts)"],
    ["• ChatGPT", "Best for planning and strategy (101 prompts)"],
    ["• Lovable", "Best for rapid prototyping (101 prompts)"],
    ["• Cursor", "Best for code-focused development (74 prompts)"],
    ["• v0", "Best for component design (55 prompts)"],
    ["", ""],
    ["TAB DESCRIPTIONS", ""],
    ["", ""],
]

# Add tab descriptions
for sheet_info in sheets_created:
    tab_name = sheet_info['name']
    count = sheet_info['count']
    description = tab_structure[tab_name]['description']
    guide_content.append([tab_name, f"{description} ({count} prompts)"])
    guide_content.append(["", ""])

guide_content.extend([
    ["TIPS FOR MAXIMUM SUCCESS", ""],
    ["", ""],
    ["✅ Always start sessions with START HERE prompts", "Prevents AI from making unwanted changes"],
    ["✅ Use Training Wheels prompts when learning", "Get detailed explanations and step-by-step guidance"],
    ["✅ Graduate to No Training Wheels for speed", "Once familiar with patterns, use concise prompts"],
    ["✅ Combine prompts from multiple tabs", "Create comprehensive solutions by mixing categories"],
    ["✅ Check tool compatibility before using", "Ensure the prompt works with your chosen AI tool"],
    ["✅ Read the Description/Notes column", "Contains valuable context and usage tips"],
    ["✅ Customize [placeholders] in prompts", "Replace bracketed text with your specific needs"],
    ["✅ Test prompts iteratively", "Start simple, then add complexity"],
    ["", ""],
    ["WORKBOOK STATISTICS", ""],
    ["", ""],
    ["Total Prompts", str(len(df))],
    ["Total Tabs", str(len(sheets_created))],
    ["Total Categories", str(df['Category'].nunique())],
    ["Last Updated", "October 21, 2025"],
    ["", ""],
    ["", ""],
    ["🎉 Happy Vibe Coding!", ""],
])

# Write guide content
for row_idx, (col1, col2) in enumerate(guide_content, 1):
    ws_guide.cell(row=row_idx, column=1, value=col1)
    ws_guide.cell(row=row_idx, column=2, value=col2)
    
    # Format headers
    if col1.startswith(('HOW TO USE', 'UNDERSTANDING', 'TOOL COMPATIBILITY', 
                       'TAB DESCRIPTIONS', 'TIPS FOR', 'WORKBOOK STATISTICS', '🎯')):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color='2C4E8C')
    
    # Format sub-headers
    if col1.startswith(('Training Wheels', 'No Training Wheels', 'Design', 'Strategy', 
                       'Most Compatible', '✅')):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=10)
    
    # Wrap text
    ws_guide.cell(row=row_idx, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws_guide.cell(row=row_idx, column=2).alignment = Alignment(vertical='top', wrap_text=True)

# Set column widths for guide
ws_guide.column_dimensions['A'].width = 40
ws_guide.column_dimensions['B'].width = 80

print("  ✓ Created Quick Reference Guide")

# Save workbook
output_path = '/home/ubuntu/vibe_coding_prompts_library_organized.xlsx'
wb.save(output_path)
print(f"\n✓ Saved organized workbook to: {output_path}")

print("\n" + "=" * 70)
print("STEP 5: GENERATING SUMMARY")
print("=" * 70)

# Create summary
summary = {
    'total_prompts': len(df),
    'total_tabs': len(sheets_created),
    'sheets': sheets_created,
    'file_path': output_path
}

# Save summary as JSON
with open('/home/ubuntu/workbook_organization_summary.json', 'w') as f:
    json.dump(summary, f, indent=2)

print(f"\n📊 REORGANIZATION COMPLETE!")
print(f"{'=' * 70}")
print(f"Total Prompts: {len(df)}")
print(f"Total Tabs Created: {len(sheets_created) + 1} (including Quick Reference Guide)")
print(f"\n📑 Tab Breakdown:")
for i, sheet_info in enumerate(sheets_created, 1):
    print(f"  {i:2d}. {sheet_info['name']:50s} - {sheet_info['count']:3d} prompts")

print(f"\n✓ All sheets formatted with:")
print(f"  • Bold blue headers")
print(f"  • Text wrapping")
print(f"  • Alternating row colors")
print(f"  • Filters enabled")
print(f"  • Optimized column widths")

print(f"\n📄 Output file: {output_path}")
print(f"📄 Debug file: ~/tab_assignments_debug.csv")
print(f"📄 Summary JSON: ~/workbook_organization_summary.json")

