import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

print("=" * 70)
print("REORGANIZING WITH IMPROVED CATEGORIZATION LOGIC")
print("=" * 70)

# Load current Excel file
df = pd.read_excel('/home/ubuntu/vibe_coding_prompts_library.xlsx')
print(f"✓ Loaded {len(df)} prompts")

def assign_to_tab_v2(row):
    """Improved categorization logic with better priority"""
    category = str(row['Category']).lower()
    use_case = str(row['Use Case']).lower()
    prompt_text = str(row['Prompt']).lower()
    prompt_type = str(row['Prompt Type'])
    
    # Priority 1: START HERE - Only true guardrail prompts
    if (('general best practices' in category or 'meta prompting' in category) and 
        any(kw in use_case for kw in ['preventing unintended', 'setting code quality', 
                                       'prompt refinement', 'documenting solutions'])):
        return "🎯 START HERE - Pre-Session Setup"
    
    # Priority 2: Award-Winning Homepage Designs (specific homepage/layout prompts)
    if any(kw in category for kw in ['visual hierarchy & layout', 'hero sections', 
                                      'award-winning design']):
        return "✨ Award-Winning Homepage Designs"
    if any(kw in use_case for kw in ['homepage', 'hero section', 'above-fold', 'f-pattern', 
                                      'z-pattern', 'attention ratio', 'awwwards']):
        return "✨ Award-Winning Homepage Designs"
    
    # Priority 3: Visual Design & Modern Aesthetics
    if any(kw in category for kw in ['visual design', 'typography', 'color psychology', 
                                      'brand identity']):
        return "🎨 Visual Design & Modern Aesthetics"
    if any(kw in use_case for kw in ['glassmorphism', 'neumorphism', 'gradient', 
                                      'kinetic typography', 'color palette', 'typography']):
        return "🎨 Visual Design & Modern Aesthetics"
    
    # Priority 4: Interactive & Immersive Elements
    if 'interactive & immersive' in category or 'animation' in category:
        return "🚀 Interactive & Immersive Elements"
    if any(kw in use_case for kw in ['parallax', 'scroll animation', '3d', 'gsap', 
                                      'lottie', 'micro-interaction', 'cursor-following', 
                                      'hover effect']):
        return "🚀 Interactive & Immersive Elements"
    
    # Priority 5: Conversion & Funnel Optimization
    if any(kw in category for kw in ['funnel design', 'funnel optimization', 
                                      'conversion optimization', 'b2b funnel', 'saas funnel']):
        return "🎯 Conversion & Funnel Optimization"
    if 'ai & automation' in category and 'funnel' in use_case:
        return "🎯 Conversion & Funnel Optimization"
    
    # Priority 6: Landing Pages & Lead Generation
    if 'landing page' in category or 'lead generation' in category:
        return "📄 Landing Pages & Lead Generation"
    if any(kw in use_case for kw in ['landing page', 'lead magnet', 'opt-in', 'lead capture']):
        return "📄 Landing Pages & Lead Generation"
    
    # Priority 7: E-commerce & Product Pages
    if 'e-commerce' in category or 'e-commerce' in use_case:
        return "💰 E-commerce & Product Pages"
    if any(kw in use_case for kw in ['product page', 'shopping cart', 'checkout', 
                                      'product showcase']):
        return "💰 E-commerce & Product Pages"
    
    # Priority 8: Dashboard & Admin Panels
    if 'dashboard' in category or 'data visualization' in category:
        return "📊 Dashboard & Admin Panels"
    if any(kw in use_case for kw in ['dashboard', 'admin panel', 'analytics dashboard', 
                                      'metrics']):
        return "📊 Dashboard & Admin Panels"
    
    # Priority 9: Social Proof & Trust Building
    if any(kw in category for kw in ['social proof', 'trust', 'testimonial', 'reputation']):
        return "📱 Social Proof & Trust Building"
    if any(kw in use_case for kw in ['testimonial', 'review', 'trust badge', 'social proof']):
        return "📱 Social Proof & Trust Building"
    
    # Priority 10: Content & Marketing
    if any(kw in category for kw in ['content marketing', 'email marketing', 'campaign']):
        return "📧 Content & Marketing"
    
    # Priority 11: API & Integration
    if 'api' in category or 'database integration' in category:
        return "🔗 API & Integration"
    if any(kw in use_case for kw in ['api endpoint', 'third-party api', 'api integration']):
        return "🔗 API & Integration"
    
    # Priority 12: Performance & Technical
    if any(kw in category for kw in ['performance', 'security', 'authentication']):
        return "⚡ Performance & Technical"
    if any(kw in use_case for kw in ['performance', 'optimization', 'security', 'authentication']):
        return "⚡ Performance & Technical"
    
    # Priority 13: SEO & Analytics
    if 'seo' in category or 'analytics & tracking' in category:
        return "🔍 SEO & Analytics"
    if any(kw in use_case for kw in ['seo', 'analytics', 'tracking', 'predictive analytics']):
        return "🔍 SEO & Analytics"
    
    # Priority 14: Testing & Debugging
    if 'testing' in category or 'debugging' in category:
        return "🧪 Testing & Debugging"
    
    # Priority 15: UI/UX Excellence (catch-all for UI/UX, navigation, accessibility, forms, etc.)
    if any(kw in category for kw in ['ui/ux', 'navigation', 'accessibility', 'responsive', 
                                      'mobile', 'form']):
        return "💎 UI/UX Excellence"
    
    # Priority 16: Feature Development (general features not fitting above)
    if 'feature development' in category:
        return "🛠️ Feature Development"
    
    # Default: Other Specialized Prompts
    return "📋 Other Specialized Prompts"

# Assign tabs
df['Tab'] = df.apply(assign_to_tab_v2, axis=1)

# Show distribution
tab_counts = df['Tab'].value_counts()
print(f"\n📊 NEW TAB DISTRIBUTION:")
for tab, count in tab_counts.items():
    percentage = (count / len(df)) * 100
    print(f"  {tab:50s} - {count:3d} prompts ({percentage:5.1f}%)")

# Tab structure with descriptions
tab_structure = {
    "🎯 START HERE - Pre-Session Setup": "Essential guardrail prompts to use before starting any vibe coding session",
    "✨ Award-Winning Homepage Designs": "Modern, futuristic homepage designs following award-winning patterns",
    "🎨 Visual Design & Modern Aesthetics": "Color psychology, typography, and cutting-edge visual design trends",
    "🚀 Interactive & Immersive Elements": "3D effects, animations, parallax, and interactive experiences",
    "💎 UI/UX Excellence": "Navigation, user experience, accessibility, and interface design",
    "🎯 Conversion & Funnel Optimization": "Sales funnels, conversion optimization, and funnel strategies",
    "📄 Landing Pages & Lead Generation": "High-converting landing pages and lead capture strategies",
    "💰 E-commerce & Product Pages": "Product showcases, shopping experiences, and e-commerce optimization",
    "📊 Dashboard & Admin Panels": "Data visualization, admin interfaces, and dashboard design",
    "📱 Social Proof & Trust Building": "Testimonials, reviews, trust badges, and credibility elements",
    "📧 Content & Marketing": "Content strategy, email marketing, and marketing campaigns",
    "🔗 API & Integration": "API integration, third-party services, and external connections",
    "⚡ Performance & Technical": "Performance optimization, security, and technical improvements",
    "🔍 SEO & Analytics": "Search optimization, tracking, and analytics implementation",
    "🧪 Testing & Debugging": "Testing strategies, debugging, and quality assurance",
    "🛠️ Feature Development": "Building specific features, components, and functionality",
    "📋 Other Specialized Prompts": "Miscellaneous prompts that don't fit other categories"
}

# Create workbook
print("\n" + "=" * 70)
print("CREATING MULTI-TAB WORKBOOK")
print("=" * 70)

wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

column_widths = {'A': 30, 'B': 60, 'C': 20, 'D': 25, 'E': 20, 'F': 40}

def format_worksheet(ws, df_tab):
    """Apply formatting to a worksheet"""
    headers = ['Use Case', 'Prompt', 'Category', 'Tool Compatibility', 'Prompt Type', 'Description/Notes']
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data
    for row_idx, (_, row) in enumerate(df_tab.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row['Use Case'])
        ws.cell(row=row_idx, column=2, value=row['Prompt'])
        ws.cell(row=row_idx, column=3, value=row['Category'])
        ws.cell(row=row_idx, column=4, value=row['Tool Compatibility'])
        ws.cell(row=row_idx, column=5, value=row['Prompt Type'])
        ws.cell(row=row_idx, column=6, value=row['Description/Notes'])
    
    # Apply formatting
    for row_idx in range(2, ws.max_row + 1):
        fill = gray_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

# Create sheets in order
tab_order = list(tab_structure.keys())
sheets_created = []

for tab_name in tab_order:
    df_tab = df[df['Tab'] == tab_name].copy()
    
    if len(df_tab) > 0:
        sheet_name = tab_name[:31].replace('/', '-')
        ws = wb.create_sheet(title=sheet_name)
        df_tab = df_tab.drop('Tab', axis=1)
        format_worksheet(ws, df_tab)
        sheets_created.append({'name': tab_name, 'sheet_name': sheet_name, 'count': len(df_tab)})
        print(f"  ✓ {sheet_name} ({len(df_tab)} prompts)")

# Create Quick Reference Guide
print("\n" + "=" * 70)
print("CREATING QUICK REFERENCE GUIDE")
print("=" * 70)

ws_guide = wb.create_sheet(title="📚 Quick Reference Guide")

guide_content = [
    ["🎯 VIBE CODING PROMPTS LIBRARY - QUICK REFERENCE GUIDE", ""],
    ["", ""],
    ["HOW TO USE THIS WORKBOOK", ""],
    ["", ""],
    ["1. Start with the 🎯 START HERE tab", "Copy and paste those prompts at the beginning of your vibe coding session"],
    ["", "This ensures the AI follows your instructions precisely and maximizes your credits"],
    ["", ""],
    ["2. Choose your design focus", "Navigate to the relevant tab based on what you're building:"],
    ["", ""],
]

# Add tab descriptions
for sheet_info in sheets_created:
    tab_name = sheet_info['name']
    count = sheet_info['count']
    description = tab_structure[tab_name]
    guide_content.append([f"  {tab_name}", f"{description} ({count} prompts)"])

guide_content.extend([
    ["", ""],
    ["3. Filter and search", "Use the filter dropdowns in each sheet to find specific:"],
    ["", "• Tool compatibility (Lovable, Replit, ChatGPT, etc.)"],
    ["", "• Prompt types (Training Wheels vs No Training Wheels)"],
    ["", "• Specific use cases"],
    ["", ""],
    ["4. Copy the prompt", "Copy the entire prompt text from column B"],
    ["", "Customize it by replacing [placeholders] with your specific needs"],
    ["", ""],
    ["5. Combine prompts", "For complex projects, combine prompts from multiple tabs"],
    ["", ""],
    ["UNDERSTANDING PROMPT TYPES", ""],
    ["", ""],
    ["Training Wheels", "Detailed, step-by-step prompts with extensive guidance"],
    ["No Training Wheels", "Concise, expert-level prompts that assume knowledge"],
    ["Design", "Visual design and UI/UX focused prompts"],
    ["Strategy", "High-level planning and strategic prompts"],
    ["", ""],
    ["TOOL COMPATIBILITY", ""],
    ["", ""],
    ["• Replit", "Best for full-stack applications"],
    ["• ChatGPT", "Best for planning and strategy"],
    ["• Lovable", "Best for rapid prototyping"],
    ["• Cursor", "Best for code-focused development"],
    ["• v0", "Best for component design"],
    ["", ""],
    ["WORKBOOK STATISTICS", ""],
    ["", ""],
    ["Total Prompts", str(len(df))],
    ["Total Tabs", str(len(sheets_created))],
    ["Last Updated", "October 21, 2025"],
    ["", ""],
    ["🎉 Happy Vibe Coding!", ""],
])

# Write guide content
for row_idx, (col1, col2) in enumerate(guide_content, 1):
    ws_guide.cell(row=row_idx, column=1, value=col1)
    ws_guide.cell(row=row_idx, column=2, value=col2)
    
    if any(col1.startswith(x) for x in ['HOW TO USE', 'UNDERSTANDING', 'TOOL COMPATIBILITY', 
                                         'WORKBOOK STATISTICS', '🎯']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color='2C4E8C')
    
    if any(col1.startswith(x) for x in ['Training Wheels', 'No Training Wheels', 'Design', 
                                         'Strategy', '•']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=10)
    
    ws_guide.cell(row=row_idx, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws_guide.cell(row=row_idx, column=2).alignment = Alignment(vertical='top', wrap_text=True)

ws_guide.column_dimensions['A'].width = 40
ws_guide.column_dimensions['B'].width = 80

print("  ✓ Created Quick Reference Guide")

# Save workbook
output_path = '/home/ubuntu/vibe_coding_prompts_library_organized.xlsx'
wb.save(output_path)

print("\n" + "=" * 70)
print("✅ REORGANIZATION COMPLETE!")
print("=" * 70)
print(f"\nTotal Prompts: {len(df)}")
print(f"Total Tabs: {len(sheets_created) + 1} (including Quick Reference Guide)")
print(f"\n📑 Final Tab Structure:")
for i, sheet_info in enumerate(sheets_created, 1):
    print(f"  {i:2d}. {sheet_info['name']:50s} - {sheet_info['count']:3d} prompts")

print(f"\n📄 Output: {output_path}")

# Save summary
summary = {
    'total_prompts': len(df),
    'total_tabs': len(sheets_created),
    'sheets': sheets_created
}

with open('/home/ubuntu/workbook_summary_v2.json', 'w') as f:
    json.dump(summary, f, indent=2)

