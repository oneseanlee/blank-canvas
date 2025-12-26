import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

print("=" * 80)
print("ADDING FINALIZATION TAB TO VIBE CODING BIBLE")
print("=" * 80)

# Load the finalization prompts
with open('/home/ubuntu/finalization_prompts.json', 'r') as f:
    finalization_prompts = json.load(f)

# Create DataFrame
df_finalization = pd.DataFrame(finalization_prompts)

# Reorder columns to match existing structure
column_order = ['Use Case', 'Prompt', 'Category', 'Tool Compatibility', 'Prompt Type', 'Description/Notes']
df_finalization = df_finalization[column_order]

print(f"\n✓ Created DataFrame with {len(df_finalization)} prompts")
print(f"  Columns: {list(df_finalization.columns)}")

# Load existing workbook
file_path = '/home/ubuntu/Uploads/Vibe Coding Bible.xlsx'
wb = load_workbook(file_path)

print(f"\n📂 Current workbook has {len(wb.sheetnames)} sheets")

# Check if finalization tab already exists
if '🏁 Project Finalization & Polish' in wb.sheetnames:
    print("  ⚠️ Tab already exists - removing old version")
    del wb['🏁 Project Finalization & Polish']

# Load a reference sheet to copy formatting
reference_sheet_name = '🎯 START HERE - Pre-Session Set'
ws_reference = wb[reference_sheet_name]

# Get formatting from reference sheet
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Create new sheet (insert after Testing & Debugging, before Feature Development)
# Find the position to insert
sheet_names = wb.sheetnames
testing_index = sheet_names.index('🧪 Testing & Debugging')
ws_new = wb.create_sheet('🏁 Project Finalization & Polish', testing_index + 1)

print(f"\n✓ Created new sheet: '🏁 Project Finalization & Polish'")

# Define column widths (matching other tabs)
column_widths = {
    'A': 30,  # Use Case
    'B': 60,  # Prompt
    'C': 20,  # Category
    'D': 25,  # Tool Compatibility
    'E': 20,  # Prompt Type
    'F': 40   # Description/Notes
}

# Write headers
headers = ['Use Case', 'Prompt', 'Category', 'Tool Compatibility', 'Prompt Type', 'Description/Notes']
for col_idx, header in enumerate(headers, 1):
    cell = ws_new.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

print(f"✓ Added headers")

# Write data
for row_idx, (_, row) in enumerate(df_finalization.iterrows(), 2):
    ws_new.cell(row=row_idx, column=1, value=row['Use Case'])
    ws_new.cell(row=row_idx, column=2, value=row['Prompt'])
    ws_new.cell(row=row_idx, column=3, value=row['Category'])
    ws_new.cell(row=row_idx, column=4, value=row['Tool Compatibility'])
    ws_new.cell(row=row_idx, column=5, value=row['Prompt Type'])
    ws_new.cell(row=row_idx, column=6, value=row['Description/Notes'])

print(f"✓ Added {len(df_finalization)} rows of data")

# Apply formatting
for row_idx in range(2, ws_new.max_row + 1):
    fill = gray_fill if row_idx % 2 == 0 else white_fill
    for col_idx in range(1, 7):
        cell = ws_new.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.alignment = Alignment(vertical='top', wrap_text=True)

print(f"✓ Applied alternating row colors and text wrapping")

# Set column widths
for col_letter, width in column_widths.items():
    ws_new.column_dimensions[col_letter].width = width

print(f"✓ Set column widths")

# Set header row height
ws_new.row_dimensions[1].height = 30

# Enable filters
ws_new.auto_filter.ref = f"A1:F{ws_new.max_row}"

print(f"✓ Enabled filters")

# Save the workbook
wb.save(file_path)
print(f"\n✅ Successfully saved updated workbook to: {file_path}")

# Print final summary
print("\n" + "=" * 80)
print("FINAL WORKBOOK SUMMARY")
print("=" * 80)

wb_check = load_workbook(file_path)
sheet_names = wb_check.sheetnames

print(f"\nTotal Sheets: {len(sheet_names)}")
print(f"\nSheet Names (in order):")
for i, name in enumerate(sheet_names, 1):
    ws_check = wb_check[name]
    row_count = ws_check.max_row - 1 if ws_check.max_row > 1 else 0
    
    if name == '🏁 Project Finalization & Polish':
        print(f"  {i}. {name} ✨ NEW! ({row_count} prompts)")
    elif '📚' in name:
        print(f"  {i}. {name} (Documentation)")
    else:
        print(f"  {i}. {name} ({row_count} prompts)")

wb_check.close()

print("\n✅ Done! Your Vibe Coding Bible now includes the finalization tab.")

