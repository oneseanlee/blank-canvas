import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

print("=" * 70)
print("FINAL REORGANIZATION - OPTIMIZED CATEGORIZATION")
print("=" * 70)

# Load current Excel file
df = pd.read_excel('/home/ubuntu/vibe_coding_prompts_library.xlsx')
print(f"✓ Loaded {len(df)} prompts")

def assign_to_tab_final(row):
    """Final optimized categorization logic"""
    category = str(row['Category']).lower()
    use_case = str(row['Use Case']).lower()
    prompt_text = str(row['Prompt']).lower()
    
    # Priority 1: START HERE - Essential guardrails only
    if (('general best practices' in category or 'meta prompting' in category) and 
        any(kw in use_case for kw in ['preventing unintended', 'setting code quality', 
                                       'prompt refinement', 'documenting solutions'])):
        return "🎯 START HERE - Pre-Session Setup"
    
    # Priority 2: Award-Winning Homepage Designs
    if any(kw in category for kw in ['visual hierarchy & layout', 'hero sections', 
                                      'award-winning design']):
        return "✨ Award-Winning Homepage Designs"
    if any(kw in use_case for kw in ['homepage', 'hero section', 'above-fold', 'f-pattern', 
                                      'z-pattern', 'attention ratio', 'awwwards', 'scroll-triggered']):
        return "✨ Award-Winning Homepage Designs"
    
    # Priority 3: Visual Design & Modern Aesthetics (expanded)
    if any(kw in category for kw in ['visual design', 'typography', 'color psychology', 
                                      'brand identity', 'visual design trends']):
        return "🎨 Visual Design & Modern Aesthetics"
    if any(kw in use_case for kw in ['glassmorphism', 'neumorphism', 'gradient', 
                                      'kinetic typography', 'color palette', 'typography',
                                      'brand identity', 'visual consistency', 'color-driven',
                                      'minimalist design']):
        return "🎨 Visual Design & Modern Aesthetics"
    
    # Priority 4: Interactive & Immersive Elements (expanded)
    if any(kw in category for kw in ['interactive & immersive', 'animation', 
                                      'interactive & navigation', 'interactive & e-commerce']):
        return "🚀 Interactive & Immersive Elements"
    if any(kw in use_case for kw in ['parallax', 'scroll animation', '3d', 'gsap', 
                                      'lottie', 'micro-interaction', 'cursor-following', 
                                      'hover effect', 'interactive map', 'gamification',
                                      'countdown timer', 'animation', 'interactive']):
        return "🚀 Interactive & Immersive Elements"
    
    # Priority 5: Conversion & Funnel Optimization (expanded to include personalization, retention)
    if any(kw in category for kw in ['funnel design', 'funnel optimization', 
                                      'conversion optimization', 'b2b funnel', 'saas funnel',
                                      'personalization', 'retention', 'engagement & retention',
                                      'growth & scaling']):
        return "🎯 Conversion & Funnel Optimization"
    if 'ai & automation' in category and 'funnel' in use_case:
        return "🎯 Conversion & Funnel Optimization"
    if any(kw in use_case for kw in ['personalization', 'retention', 're-engagement', 
                                      'viral referral', 'partner funnel', 'affiliate funnel']):
        return "🎯 Conversion & Funnel Optimization"
    
    # Priority 6: Landing Pages & Lead Generation (expanded to include conversion elements)
    if any(kw in category for kw in ['landing page', 'lead generation', 'conversion & urgency',
                                      'conversion & comparison', 'conversion-focused']):
        return "📄 Landing Pages & Lead Generation"
    if any(kw in use_case for kw in ['landing page', 'lead magnet', 'opt-in', 'lead capture',
                                      'countdown timer', 'comparison table', 'limited offer']):
        return "📄 Landing Pages & Lead Generation"
    
    # Priority 7: E-commerce & Product Pages
    if 'e-commerce' in category or 'e-commerce' in use_case:
        return "💰 E-commerce & Product Pages"
    if any(kw in use_case for kw in ['product page', 'shopping', 'checkout', 
                                      'product showcase', 'voice search', 'conversational commerce']):
        return "💰 E-commerce & Product Pages"
    
    # Priority 8: Dashboard & Admin Panels
    if 'dashboard' in category or 'data visualization' in category:
        return "📊 Dashboard & Admin Panels"
    if any(kw in use_case for kw in ['dashboard', 'admin panel', 'analytics dashboard', 
                                      'metrics', 'data visualization']):
        return "📊 Dashboard & Admin Panels"
    
    # Priority 9: Social Proof & Trust Building (expanded to include about/contact sections)
    if any(kw in category for kw in ['social proof', 'trust', 'testimonial', 'reputation',
                                      'about & team', 'contact']):
        return "📱 Social Proof & Trust Building"
    if any(kw in use_case for kw in ['testimonial', 'review', 'trust badge', 'social proof',
                                      'team section', 'about section', 'contact section',
                                      'faq section']):
        return "📱 Social Proof & Trust Building"
    
    # Priority 10: Content & Marketing (expanded to include subscription)
    if any(kw in category for kw in ['content marketing', 'email marketing', 'campaign',
                                      'subscription', 'local marketing', 'promotion']):
        return "📧 Content & Marketing"
    if any(kw in use_case for kw in ['content strategy', 'email', 'marketing', 'campaign', 
                                      'newsletter', 'subscription', 'local business']):
        return "📧 Content & Marketing"
    
    # Priority 11: API & Integration
    if 'api' in category or 'database integration' in category:
        return "🔗 API & Integration"
    if any(kw in use_case for kw in ['api endpoint', 'third-party api', 'api integration']):
        return "🔗 API & Integration"
    
    # Priority 12: Performance & Technical (expanded to include emerging tech)
    if any(kw in category for kw in ['performance', 'security', 'authentication', 
                                      'emerging technologies']):
        return "⚡ Performance & Technical"
    if any(kw in use_case for kw in ['performance', 'optimization', 'security', 
                                      'authentication', 'refactoring']):
        return "⚡ Performance & Technical"
    
    # Priority 13: SEO & Analytics
    if 'seo' in category or 'analytics & tracking' in category:
        return "🔍 SEO & Analytics"
    if any(kw in use_case for kw in ['seo', 'analytics', 'tracking', 'predictive analytics']):
        return "🔍 SEO & Analytics"
    
    # Priority 14: Testing & Debugging
    if 'testing' in category or 'debugging' in category:
        return "🧪 Testing & Debugging"
    if any(kw in use_case for kw in ['testing', 'debugging', 'overview before implementation']):
        return "🧪 Testing & Debugging"
    
    # Priority 15: UI/UX Excellence
    if any(kw in category for kw in ['ui/ux', 'navigation', 'accessibility', 'responsive', 
                                      'mobile', 'form', 'search & discovery', 'content & ux']):
        return "💎 UI/UX Excellence"
    
    # Priority 16: Legal & Compliance
    if 'legal' in category or 'compliance' in category:
        return "⚖️ Legal & Compliance"
    
    # Priority 17: Feature Development
    if 'feature development' in category:
        return "🛠️ Feature Development"
    
    # Default: Other Specialized Prompts
    return "📋 Other Specialized Prompts"

# Assign tabs
df['Tab'] = df.apply(assign_to_tab_final, axis=1)

# Show distribution
tab_counts = df['Tab'].value_counts()
print(f"\n📊 FINAL TAB DISTRIBUTION:")
total = len(df)
for tab, count in sorted(tab_counts.items(), key=lambda x: -x[1]):
    percentage = (count / total) * 100
    print(f"  {tab:50s} - {count:3d} prompts ({percentage:5.1f}%)")

# Tab structure
tab_structure = {
    "🎯 START HERE - Pre-Session Setup": "Essential guardrail prompts to use before starting any vibe coding session",
    "✨ Award-Winning Homepage Designs": "Modern, futuristic homepage designs following award-winning patterns",
    "🎨 Visual Design & Modern Aesthetics": "Color psychology, typography, brand identity, and cutting-edge visual design trends",
    "🚀 Interactive & Immersive Elements": "3D effects, animations, parallax, interactive experiences, and gamification",
    "💎 UI/UX Excellence": "Navigation, user experience, accessibility, forms, and interface design",
    "🎯 Conversion & Funnel Optimization": "Sales funnels, conversion optimization, personalization, and growth strategies",
    "📄 Landing Pages & Lead Generation": "High-converting landing pages, lead capture, and conversion-focused elements",
    "💰 E-commerce & Product Pages": "Product showcases, shopping experiences, and e-commerce optimization",
    "📊 Dashboard & Admin Panels": "Data visualization, admin interfaces, and dashboard design",
    "📱 Social Proof & Trust Building": "Testimonials, reviews, trust elements, about/team sections, and contact pages",
    "📧 Content & Marketing": "Content strategy, email marketing, campaigns, and subscription models",
    "🔗 API & Integration": "API integration, third-party services, and external connections",
    "⚡ Performance & Technical": "Performance optimization, security, refactoring, and emerging technologies",
    "🔍 SEO & Analytics": "Search optimization, tracking, and analytics implementation",
    "🧪 Testing & Debugging": "Testing strategies, debugging, and quality assurance",
    "⚖️ Legal & Compliance": "Legal requirements, privacy policies, and compliance",
    "🛠️ Feature Development": "Building specific features, components, and functionality",
    "📋 Other Specialized Prompts": "Miscellaneous specialized prompts"
}

# Create workbook
print("\n" + "=" * 70)
print("CREATING FINAL MULTI-TAB WORKBOOK")
print("=" * 70)

wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
column_widths = {'A': 30, 'B': 60, 'C': 20, 'D': 25, 'E': 20, 'F': 40}

def format_worksheet(ws, df_tab):
    headers = ['Use Case', 'Prompt', 'Category', 'Tool Compatibility', 'Prompt Type', 'Description/Notes']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, (_, row) in enumerate(df_tab.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row['Use Case'])
        ws.cell(row=row_idx, column=2, value=row['Prompt'])
        ws.cell(row=row_idx, column=3, value=row['Category'])
        ws.cell(row=row_idx, column=4, value=row['Tool Compatibility'])
        ws.cell(row=row_idx, column=5, value=row['Prompt Type'])
        ws.cell(row=row_idx, column=6, value=row['Description/Notes'])
    
    for row_idx in range(2, ws.max_row + 1):
        fill = gray_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

tab_order = list(tab_structure.keys())
sheets_created = []

for tab_name in tab_order:
    df_tab = df[df['Tab'] == tab_name].copy()
    if len(df_tab) > 0:
        sheet_name = tab_name[:31].replace('/', '-')
        ws = wb.create_sheet(title=sheet_name)
        df_tab = df_tab.drop('Tab', axis=1)
        format_worksheet(ws, df_tab)
        sheets_created.append({'name': tab_name, 'count': len(df_tab)})
        print(f"  ✓ {sheet_name} ({len(df_tab)} prompts)")

# Create Quick Reference Guide
print("\n" + "=" * 70)
print("CREATING COMPREHENSIVE QUICK REFERENCE GUIDE")
print("=" * 70)

ws_guide = wb.create_sheet(title="📚 Quick Reference Guide")

guide_content = [
    ["🎯 VIBE CODING PROMPTS LIBRARY - QUICK REFERENCE GUIDE", ""],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["HOW TO USE THIS WORKBOOK", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["STEP 1: Start Your Session Right", ""],
    ["   📍 Go to: 🎯 START HERE - Pre-Session Setup", "Always begin by copying these essential guardrail prompts"],
    ["", "These prompts ensure the AI:"],
    ["", "   • Makes only the changes you request"],
    ["", "   • Maintains code quality and consistency"],
    ["", "   • Maximizes your credits by reducing back-and-forth"],
    ["", "   • Follows modern development best practices"],
    ["", ""],
    ["STEP 2: Choose Your Design Category", ""],
    ["", "Navigate to the tab that matches your project:"],
    ["", ""],
]

# Add all tab descriptions
for sheet_info in sheets_created:
    tab_name = sheet_info['name']
    count = sheet_info['count']
    description = tab_structure.get(tab_name, "")
    guide_content.append([f"   {tab_name}", f"{description} ({count} prompts)"])

guide_content.extend([
    ["", ""],
    ["", ""],
    ["STEP 3: Filter and Find", ""],
    ["", "Use Excel's built-in filters (click the dropdown arrows in the header row):"],
    ["", ""],
    ["   🔍 Filter by Tool Compatibility", "Find prompts that work with your AI tool:"],
    ["", "   • Lovable - Best for rapid prototyping and visual designs"],
    ["", "   • Replit - Best for full-stack applications"],
    ["", "   • ChatGPT - Best for planning, strategy, and content"],
    ["", "   • Cursor - Best for code-focused development"],
    ["", "   • v0 - Best for component design and UI"],
    ["", ""],
    ["   📝 Filter by Prompt Type", "Choose the right level of detail:"],
    ["", "   • Training Wheels - Detailed, step-by-step guidance (great for learning)"],
    ["", "   • No Training Wheels - Concise, expert-level prompts (for speed)"],
    ["", "   • Design - Visual design and UI/UX focused"],
    ["", "   • Strategy - High-level planning and business strategy"],
    ["", ""],
    ["STEP 4: Customize Your Prompt", ""],
    ["", "1. Copy the entire prompt text from Column B"],
    ["", "2. Look for [placeholders] in square brackets"],
    ["", "3. Replace them with your specific information:"],
    ["", "   Example: [product/service] → \"AI-powered analytics platform\""],
    ["", "   Example: [target audience] → \"SaaS founders and product managers\""],
    ["", "   Example: [brand colors] → \"#3B82F6 (blue), #10B981 (green)\""],
    ["", ""],
    ["STEP 5: Combine for Complex Projects", ""],
    ["", "For comprehensive solutions, combine prompts from multiple tabs:"],
    ["", ""],
    ["   Example: Building a High-Converting SaaS Landing Page"],
    ["", "   1. START HERE prompts (guardrails)"],
    ["", "   2. ✨ Award-Winning Homepage Designs (layout and hero section)"],
    ["", "   3. 🎨 Visual Design & Modern Aesthetics (colors and typography)"],
    ["", "   4. 🚀 Interactive & Immersive Elements (animations)"],
    ["", "   5. 📄 Landing Pages & Lead Generation (conversion elements)"],
    ["", "   6. 📱 Social Proof & Trust Building (testimonials)"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["UNDERSTANDING PROMPT TYPES", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎓 Training Wheels", "Best for: Learning, complex projects, detailed control"],
    ["", "Characteristics:"],
    ["", "   • Step-by-step instructions"],
    ["", "   • Extensive explanations and context"],
    ["", "   • Multiple examples provided"],
    ["", "   • Detailed specifications"],
    ["", "When to use: Starting out, learning new patterns, complex requirements"],
    ["", ""],
    ["⚡ No Training Wheels", "Best for: Speed, experienced users, simple tasks"],
    ["", "Characteristics:"],
    ["", "   • Concise and direct"],
    ["", "   • Assumes prior knowledge"],
    ["", "   • Focuses on end result"],
    ["", "   • Minimal explanation"],
    ["", "When to use: Familiar patterns, quick iterations, prototyping"],
    ["", ""],
    ["🎨 Design", "Best for: Visual design, UI/UX, aesthetics"],
    ["", "Characteristics:"],
    ["", "   • Focus on visual elements"],
    ["", "   • Color, typography, layout details"],
    ["", "   • User experience considerations"],
    ["", "   • Design principles and patterns"],
    ["", "When to use: Creating interfaces, improving visuals, design systems"],
    ["", ""],
    ["📊 Strategy", "Best for: Business planning, high-level decisions"],
    ["", "Characteristics:"],
    ["", "   • Business-focused outcomes"],
    ["", "   • Strategic frameworks"],
    ["", "   • Planning and architecture"],
    ["", "   • ROI and metrics-oriented"],
    ["", "When to use: Project planning, funnel design, business strategy"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["PRO TIPS FOR MAXIMUM SUCCESS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["✅ ALWAYS start with START HERE prompts", "Save credits and prevent unwanted changes"],
    ["", ""],
    ["✅ Start with Training Wheels, graduate to No Training Wheels", "Build understanding first, then speed up"],
    ["", ""],
    ["✅ Read the Description/Notes column", "Contains valuable context, tips, and best practices"],
    ["", ""],
    ["✅ Test iteratively", "Start with core functionality, then add complexity"],
    ["", ""],
    ["✅ Combine design prompts", "Layer multiple design elements for award-winning results"],
    ["", ""],
    ["✅ Customize for your brand", "Replace generic terms with your specific brand voice"],
    ["", ""],
    ["✅ Save your successful combinations", "Keep a document of prompt combinations that work well"],
    ["", ""],
    ["✅ Check tool compatibility first", "Ensure the prompt works with your chosen AI tool"],
    ["", ""],
    ["✅ Use specific examples", "The more specific your [placeholders], the better the output"],
    ["", ""],
    ["✅ Reference existing designs", "Include links to designs you admire for visual reference"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["COMMON USE CASES & TAB COMBINATIONS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎯 Building a SaaS Landing Page", "Recommended tabs:"],
    ["", "   1. 🎯 START HERE"],
    ["", "   2. ✨ Award-Winning Homepage Designs"],
    ["", "   3. 📄 Landing Pages & Lead Generation"],
    ["", "   4. 💎 UI/UX Excellence"],
    ["", "   5. 📱 Social Proof & Trust Building"],
    ["", ""],
    ["🛒 Creating an E-commerce Store", "Recommended tabs:"],
    ["", "   1. 🎯 START HERE"],
    ["", "   2. 💰 E-commerce & Product Pages"],
    ["", "   3. 💎 UI/UX Excellence"],
    ["", "   4. 🚀 Interactive & Immersive Elements"],
    ["", "   5. ⚡ Performance & Technical"],
    ["", ""],
    ["📊 Building a Dashboard", "Recommended tabs:"],
    ["", "   1. 🎯 START HERE"],
    ["", "   2. 📊 Dashboard & Admin Panels"],
    ["", "   3. 💎 UI/UX Excellence"],
    ["", "   4. ⚡ Performance & Technical"],
    ["", ""],
    ["🎨 Portfolio/Agency Website", "Recommended tabs:"],
    ["", "   1. 🎯 START HERE"],
    ["", "   2. ✨ Award-Winning Homepage Designs"],
    ["", "   3. 🎨 Visual Design & Modern Aesthetics"],
    ["", "   4. 🚀 Interactive & Immersive Elements"],
    ["", "   5. 📱 Social Proof & Trust Building"],
    ["", ""],
    ["💰 Sales Funnel", "Recommended tabs:"],
    ["", "   1. 🎯 START HERE"],
    ["", "   2. 🎯 Conversion & Funnel Optimization"],
    ["", "   3. 📄 Landing Pages & Lead Generation"],
    ["", "   4. 📧 Content & Marketing"],
    ["", "   5. 🔍 SEO & Analytics"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["WORKBOOK STATISTICS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["Total Prompts", str(len(df))],
    ["Total Tabs", str(len(sheets_created))],
    ["Total Categories", str(df['Category'].nunique())],
    ["Last Updated", "October 21, 2025"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎉 Ready to create something amazing?", "Start with the 🎯 START HERE tab and let's go!"],
    ["", ""],
])

# Write guide content with formatting
for row_idx, (col1, col2) in enumerate(guide_content, 1):
    ws_guide.cell(row=row_idx, column=1, value=col1)
    ws_guide.cell(row=row_idx, column=2, value=col2)
    
    # Format section headers
    if any(col1.startswith(x) for x in ['HOW TO USE', 'UNDERSTANDING', 'PRO TIPS', 
                                         'COMMON USE', 'WORKBOOK STATISTICS', '🎯 VIBE',
                                         '═══']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color='2C4E8C')
    
    # Format step headers
    if any(col1.startswith(x) for x in ['STEP 1', 'STEP 2', 'STEP 3', 'STEP 4', 'STEP 5']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=11, color='1F4788')
    
    # Format sub-headers
    if any(col1.startswith(x) for x in ['🎓', '⚡', '🎨', '📊', '✅', '🎯', '🛒', '🎨', '💰']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=10)
    
    ws_guide.cell(row=row_idx, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws_guide.cell(row=row_idx, column=2).alignment = Alignment(vertical='top', wrap_text=True)

ws_guide.column_dimensions['A'].width = 45
ws_guide.column_dimensions['B'].width = 75

print("  ✓ Created comprehensive Quick Reference Guide")

# Save workbook
output_path = '/home/ubuntu/vibe_coding_prompts_library_organized.xlsx'
wb.save(output_path)

print("\n" + "=" * 70)
print("✅ ✅ ✅  REORGANIZATION COMPLETE!  ✅ ✅ ✅")
print("=" * 70)
print(f"\n📊 Total Prompts: {len(df)}")
print(f"📑 Total Tabs: {len(sheets_created) + 1} (including Quick Reference Guide)")
print(f"\n🎯 Final Tab Structure:")
for i, sheet_info in enumerate(sheets_created, 1):
    print(f"  {i:2d}. {sheet_info['name']:50s} - {sheet_info['count']:3d} prompts")

print(f"\n📄 Output File: {output_path}")
print(f"\n💡 Key Improvements:")
print(f"  ✓ Design prompts consolidated and prioritized")
print(f"  ✓ Conversion and marketing prompts grouped strategically")
print(f"  ✓ 'Other' category minimized through better categorization")
print(f"  ✓ Comprehensive Quick Reference Guide with use cases")
print(f"  ✓ Professional formatting applied to all sheets")

# Save summary
summary = {
    'total_prompts': len(df),
    'total_tabs': len(sheets_created),
    'sheets': sheets_created,
    'file_path': output_path
}

with open('/home/ubuntu/final_workbook_summary.json', 'w') as f:
    json.dump(summary, f, indent=2)

print(f"\n✓ Summary saved to: ~/final_workbook_summary.json")

