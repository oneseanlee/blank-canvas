import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the JSON data
with open('/home/ubuntu/vibe_coding_prompts_data.json', 'r') as f:
    data = json.load(f)

# Extract prompts data
prompts = data['prompts']

# Convert to DataFrame
df_data = []
for prompt in prompts:
    # Convert tool_compatibility list to comma-separated string
    tools = ', '.join(prompt['tool_compatibility']) if isinstance(prompt['tool_compatibility'], list) else prompt['tool_compatibility']
    
    df_data.append({
        'Use Case': prompt['use_case'],
        'Prompt': prompt['prompt_text'],
        'Category': prompt['category'],
        'Tool Compatibility': tools,
        'Prompt Type': prompt['prompt_type'],
        'Description/Notes': prompt['description']
    })

# Create DataFrame
df = pd.DataFrame(df_data)

# Sort by Category to group similar prompts together
df = df.sort_values(by='Category', ascending=True).reset_index(drop=True)

# Save to Excel
output_file = '/home/ubuntu/vibe_coding_prompts_library.xlsx'
df.to_excel(output_file, index=False, sheet_name='Vibe Coding Prompts')

# Now apply formatting using openpyxl
wb = load_workbook(output_file)
ws = wb.active

# Define styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
cell_font = Font(name='Calibri', size=10)
alignment_wrap = Alignment(wrap_text=True, vertical='top')
alignment_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
border_thin = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Alternating row colors
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
fill_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# Apply header formatting
for col_num, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = border_thin

# Set column widths (more generous for readability)
column_widths = {
    'A': 35,  # Use Case
    'B': 80,  # Prompt (widest)
    'C': 25,  # Category
    'D': 30,  # Tool Compatibility
    'E': 20,  # Prompt Type
    'F': 45   # Description/Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Apply cell formatting and alternating row colors
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    # Determine fill color (alternating)
    if row_num % 2 == 0:
        fill = fill_gray
    else:
        fill = fill_white
    
    for cell in row:
        cell.font = cell_font
        cell.alignment = alignment_wrap
        cell.border = border_thin
        cell.fill = fill

# Add AutoFilter to header row
ws.auto_filter.ref = f'A1:F{ws.max_row}'

# Freeze the header row
ws.freeze_panes = 'A2'

# Save the workbook
wb.save(output_file)

print(f"✅ Excel file created successfully: {output_file}")
print(f"📊 Total prompts included: {len(df)}")
print(f"📂 Categories: {df['Category'].nunique()}")
print(f"\nCategory breakdown:")
category_counts = df['Category'].value_counts().sort_index()
for category, count in category_counts.items():
    print(f"  • {category}: {count} prompts")
