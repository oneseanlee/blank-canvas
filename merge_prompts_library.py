import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher

def similarity(a, b):
    """Calculate similarity ratio between two strings"""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def load_json_file(filepath):
    """Load JSON file and return data"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Handle different JSON structures
        if isinstance(data, dict) and 'prompts' in data:
            # Extract prompts array from dict structure
            prompts = data['prompts']
            print(f"✓ Loaded {filepath}: {len(prompts)} prompts (dict format)")
            return prompts
        elif isinstance(data, list):
            print(f"✓ Loaded {filepath}: {len(data)} prompts (list format)")
            return data
        else:
            print(f"✗ Unexpected format in {filepath}")
            return []
    except Exception as e:
        print(f"✗ Error loading {filepath}: {e}")
        return []

def normalize_prompt(prompt):
    """Normalize prompt structure to have consistent field names"""
    normalized = {}
    
    # Map various field names to standard names
    normalized['use_case'] = prompt.get('use_case', '')
    normalized['prompt'] = prompt.get('prompt', prompt.get('prompt_text', ''))
    normalized['category'] = prompt.get('category', '')
    normalized['tool_compatibility'] = prompt.get('tool_compatibility', [])
    normalized['prompt_type'] = prompt.get('prompt_type', '')
    normalized['description'] = prompt.get('description', '')
    
    return normalized

def remove_duplicates(prompts, similarity_threshold=0.85):
    """Remove duplicate prompts based on text similarity"""
    unique_prompts = []
    duplicates_removed = 0
    
    for prompt in prompts:
        is_duplicate = False
        prompt_text = prompt.get('prompt', '')
        
        for unique_prompt in unique_prompts:
            unique_text = unique_prompt.get('prompt', '')
            if similarity(prompt_text, unique_text) >= similarity_threshold:
                is_duplicate = True
                duplicates_removed += 1
                break
        
        if not is_duplicate:
            unique_prompts.append(prompt)
    
    return unique_prompts, duplicates_removed

# Load all three JSON files
print("=" * 60)
print("LOADING JSON FILES")
print("=" * 60)

file1 = load_json_file('/home/ubuntu/vibe_coding_prompts_data.json')
file2 = load_json_file('/home/ubuntu/high_converting_funnels_prompts.json')
file3 = load_json_file('/home/ubuntu/award_winning_design_prompts.json')

# Normalize all prompts to have consistent field names
print("\n" + "=" * 60)
print("NORMALIZING DATA STRUCTURES")
print("=" * 60)
file1_normalized = [normalize_prompt(p) for p in file1]
file2_normalized = [normalize_prompt(p) for p in file2]
file3_normalized = [normalize_prompt(p) for p in file3]
print(f"✓ Normalized all prompts to consistent structure")

# Combine all prompts
print("\n" + "=" * 60)
print("MERGING DATA")
print("=" * 60)
all_prompts = file1_normalized + file2_normalized + file3_normalized
print(f"Total prompts before deduplication: {len(all_prompts)}")

# Remove duplicates
unique_prompts, duplicates = remove_duplicates(all_prompts)
print(f"Duplicates removed: {duplicates}")
print(f"Total unique prompts: {len(unique_prompts)}")

# Sort by category for better organization
unique_prompts_sorted = sorted(unique_prompts, key=lambda x: x.get('category', ''))

# Save merged data back to JSON
with open('/home/ubuntu/vibe_coding_prompts_data.json', 'w', encoding='utf-8') as f:
    json.dump(unique_prompts_sorted, f, indent=2, ensure_ascii=False)
print(f"\n✓ Saved merged data to ~/vibe_coding_prompts_data.json")

# Create DataFrame for Excel
df = pd.DataFrame(unique_prompts_sorted)

# Convert tool_compatibility list to string
df['tool_compatibility'] = df['tool_compatibility'].apply(lambda x: ', '.join(x) if isinstance(x, list) else str(x))

# Ensure all required columns exist and reorder them
column_order = ['use_case', 'prompt', 'category', 'tool_compatibility', 'prompt_type', 'description']
for col in column_order:
    if col not in df.columns:
        df[col] = ''

df = df[column_order]

# Rename columns for Excel
df.columns = ['Use Case', 'Prompt', 'Category', 'Tool Compatibility', 'Prompt Type', 'Description/Notes']

# Save to Excel
excel_path = '/home/ubuntu/vibe_coding_prompts_library.xlsx'
df.to_excel(excel_path, index=False, sheet_name='Vibe Coding Prompts')

# Apply formatting
wb = load_workbook(excel_path)
ws = wb.active

# Define styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Format headers (row 1)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Apply alternating row colors and text wrapping
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    fill = gray_fill if row_idx % 2 == 0 else white_fill
    for cell in row:
        cell.fill = fill
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 30,  # Use Case
    'B': 60,  # Prompt
    'C': 20,  # Category
    'D': 25,  # Tool Compatibility
    'E': 20,  # Prompt Type
    'F': 40   # Description/Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Set row heights
ws.row_dimensions[1].height = 30  # Header row
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = None  # Auto-adjust based on content

# Enable filters on header row
ws.auto_filter.ref = ws.dimensions

# Save formatted Excel file
wb.save(excel_path)
print(f"✓ Saved formatted Excel file to ~/vibe_coding_prompts_library.xlsx")

# Generate summary statistics
print("\n" + "=" * 60)
print("SUMMARY STATISTICS")
print("=" * 60)
print(f"Total prompts in final library: {len(unique_prompts_sorted)}")
print(f"Duplicates removed: {duplicates}")
print(f"\nBreakdown by category:")

category_counts = df['Category'].value_counts().sort_index()
for category, count in category_counts.items():
    print(f"  • {category}: {count} prompts")

print("\n" + "=" * 60)
print("✓ MERGE COMPLETE!")
print("=" * 60)
print("\nFiles updated:")
print("  1. ~/vibe_coding_prompts_data.json (merged JSON)")
print("  2. ~/vibe_coding_prompts_library.xlsx (formatted Excel)")
print("\nFormatting applied:")
print("  ✓ Bold headers with blue background")
print("  ✓ Text wrapping for long content")
print("  ✓ Alternating row colors (gray/white)")
print("  ✓ Filters enabled on header row")
print("  ✓ Auto-adjusted column widths")
print("  ✓ Professional layout ready for use")

