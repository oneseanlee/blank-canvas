import pandas as pd
from openpyxl import load_workbook

# Load workbook
wb = load_workbook('/home/ubuntu/vibe_coding_prompts_library_organized.xlsx')

# Get all sheets
all_sheets = wb.sheetnames
print("All sheets in workbook:")
total_prompts = 0
for sheet_name in all_sheets:
    if sheet_name != "📚 Quick Reference Guide":
        ws = wb[sheet_name]
        count = ws.max_row - 1
        total_prompts += count
        print(f"  {sheet_name:50s} - {count:3d} prompts")

print(f"\n✅ Total prompts across all tabs: {total_prompts}")
print(f"✅ Total tabs: {len(all_sheets)}")

wb.close()

# Validate against original
df_original = pd.read_excel('/home/ubuntu/vibe_coding_prompts_library.xlsx')
print(f"\n📊 Original library had: {len(df_original)} prompts")
print(f"📊 Reorganized workbook has: {total_prompts} prompts")

if total_prompts == len(df_original):
    print("\n✅ ✅ ✅ ALL PROMPTS ACCOUNTED FOR!")
else:
    print(f"\n⚠️ Mismatch: {len(df_original) - total_prompts} prompts difference")

