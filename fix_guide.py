import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import json

# Load the workbook
wb = load_workbook('/home/ubuntu/vibe_coding_prompts_library_organized.xlsx')

# Get sheet names
sheet_names = wb.sheetnames
print(f"Current sheets: {sheet_names}")

# Load summary for sheet info
with open('/home/ubuntu/final_workbook_summary.json', 'r') as f:
    summary = json.load(f)

sheets_created = summary['sheets']

# Create Quick Reference Guide
if "📚 Quick Reference Guide" in wb.sheetnames:
    del wb["📚 Quick Reference Guide"]

ws_guide = wb.create_sheet(title="📚 Quick Reference Guide")

guide_content = [
    ["🎯 VIBE CODING PROMPTS LIBRARY - QUICK REFERENCE GUIDE", ""],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["HOW TO USE THIS WORKBOOK", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["STEP 1: Start Your Session Right", ""],
    ["   📍 Go to: 🎯 START HERE - Pre-Session Setup", "Always begin by copying these essential guardrail prompts"],
    ["", "These prompts ensure the AI:"],
    ["", "   • Makes only the changes you request"],
    ["", "   • Maintains code quality and consistency"],
    ["", "   • Maximizes your credits by reducing back-and-forth"],
    ["", "   • Follows modern development best practices"],
    ["", ""],
    ["STEP 2: Choose Your Design Category", ""],
    ["", "Navigate to the tab that matches your project:"],
    ["", ""],
]

# Add all tab descriptions
for sheet_info in sheets_created:
    tab_name = sheet_info['name']
    count = sheet_info['count']
    guide_content.append([f"   {tab_name}", f"({count} prompts)"])

guide_content.extend([
    ["", ""],
    ["STEP 3: Filter and Find Your Prompts", ""],
    ["", "Use Excel's built-in filters (click dropdown arrows in header):"],
    ["", "   • Filter by Tool Compatibility (Lovable, Replit, ChatGPT, etc.)"],
    ["", "   • Filter by Prompt Type (Training Wheels, No Training Wheels, etc.)"],
    ["", "   • Search for specific keywords in Use Case or Description"],
    ["", ""],
    ["STEP 4: Customize Your Prompt", ""],
    ["", "1. Copy the entire prompt text from Column B"],
    ["", "2. Replace [placeholders] with your specific information"],
    ["", "3. Paste into your AI tool and iterate"],
    ["", ""],
    ["STEP 5: Combine for Complex Projects", ""],
    ["", "Layer multiple prompts from different tabs for best results"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["UNDERSTANDING PROMPT TYPES", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎓 Training Wheels", "Detailed, step-by-step guidance - Best for learning"],
    ["⚡ No Training Wheels", "Concise, expert-level prompts - Best for speed"],
    ["🎨 Design", "Visual design and UI/UX focused"],
    ["📊 Strategy", "High-level planning and business strategy"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["PRO TIPS FOR SUCCESS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["✅ ALWAYS start with START HERE prompts", "Prevents unwanted changes and saves credits"],
    ["✅ Read the Description/Notes column", "Contains valuable context and best practices"],
    ["✅ Start simple, then add complexity", "Test core functionality before adding features"],
    ["✅ Combine design prompts", "Layer multiple elements for award-winning results"],
    ["✅ Check tool compatibility", "Ensure prompt works with your chosen AI tool"],
    ["✅ Save successful combinations", "Keep a document of what works well"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["COMMON USE CASES", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["🎯 SaaS Landing Page", "START HERE + Award-Winning + Landing Pages + Social Proof"],
    ["🛒 E-commerce Store", "START HERE + E-commerce + UI/UX + Interactive + Performance"],
    ["📊 Dashboard", "START HERE + Dashboard + UI/UX + Performance"],
    ["🎨 Portfolio Website", "START HERE + Award-Winning + Visual Design + Interactive"],
    ["💰 Sales Funnel", "START HERE + Conversion + Landing Pages + Content + SEO"],
    ["", ""],
    ["═══════════════════════════════════════════", ""],
    ["WORKBOOK STATISTICS", ""],
    ["═══════════════════════════════════════════", ""],
    ["", ""],
    ["Total Prompts", str(summary['total_prompts'])],
    ["Total Tabs", str(summary['total_tabs'])],
    ["Last Updated", "October 21, 2025"],
    ["", ""],
    ["", ""],
    ["🎉 Ready to create something amazing?", "Start with 🎯 START HERE and let's build!"],
])

# Write guide content
for row_idx, (col1, col2) in enumerate(guide_content, 1):
    ws_guide.cell(row=row_idx, column=1, value=col1)
    ws_guide.cell(row=row_idx, column=2, value=col2)
    
    # Format headers
    if any(col1.startswith(x) for x in ['HOW TO USE', 'UNDERSTANDING', 'PRO TIPS', 
                                         'COMMON USE', 'WORKBOOK STATISTICS', '🎯 VIBE',
                                         '═══']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color='2C4E8C')
    
    # Format steps
    if any(col1.startswith(x) for x in ['STEP 1', 'STEP 2', 'STEP 3', 'STEP 4', 'STEP 5']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=11, color='1F4788')
    
    # Format items
    if any(col1.startswith(x) for x in ['🎓', '⚡', '🎨', '📊', '✅', '🎯', '🛒', '💰']):
        ws_guide.cell(row=row_idx, column=1).font = Font(bold=True, size=10)
    
    ws_guide.cell(row=row_idx, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws_guide.cell(row=row_idx, column=2).alignment = Alignment(vertical='top', wrap_text=True)

ws_guide.column_dimensions['A'].width = 45
ws_guide.column_dimensions['B'].width = 75

# Save workbook
wb.save('/home/ubuntu/vibe_coding_prompts_library_organized.xlsx')
print("\n✅ Quick Reference Guide created successfully!")

